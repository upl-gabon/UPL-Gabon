#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UPL — États financiers 2022-2025 (outil de reconstitution)
2022-2024 : estimatifs | 2025 : reconstitué (calé sur le dossier ECOBANK 260 M)
Cellules JAUNES = à saisir / ajuster avec les relevés UGB. Cellules blanches = calcul auto.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "UPL_Etats_Financiers_2022-2025.xlsx")

NAVY, GOLD, GOLD_DK = "0B2A5B", "C9A227", "A8871F"
SILK, PALE, LINE = "F7F0DC", "FCF8EE", "D8C58A"
INPUT = "FFF2CC"   # cellules à saisir

F_T = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_H = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
F_S = Font(name="Calibri", size=10.5, bold=True, color=NAVY)
F_B = Font(name="Calibri", size=10.5, color="33383F")
F_BB = Font(name="Calibri", size=10.5, bold=True, color=NAVY)
F_N = Font(name="Calibri", size=9, italic=True, color="6E6455")
FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_S = PatternFill("solid", fgColor=SILK)
FILL_Z = PatternFill("solid", fgColor=PALE)
FILL_I = PatternFill("solid", fgColor=INPUT)
THIN = Border(*[Side(style="thin", color=LINE)] * 4)
NUM = "#,##0"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def head(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, title).font = F_T
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, sub).font = F_N
    ws.row_dimensions[1].height = 22


def hrow(ws, r, values):
    for j, v in enumerate(values, 1):
        c = ws.cell(r, j, v); c.font = F_H; c.fill = FILL_H; c.alignment = CENTER; c.border = THIN


def row(ws, r, values, bold=False, fill=None, num=True, inputs=()):
    for j, v in enumerate(values, 1):
        c = ws.cell(r, j)
        c.value = v
        c.font = F_BB if bold else F_B
        c.border = THIN
        if j == 1:
            c.alignment = LEFT
        elif isinstance(v, str) and not v.startswith("="):
            c.alignment = CENTER
        elif isinstance(v, str):
            c.alignment = RIGHT; c.number_format = NUM
        else:
            c.alignment = RIGHT
            if num and isinstance(v, (int, float)):
                c.number_format = NUM
        if j - 1 in inputs:
            c.fill = FILL_I
        elif fill:
            c.fill = fill


wb = Workbook()

# ============================================================ LISEZ-MOI
ws = wb.active
ws.title = "Lisez-moi"
ws.sheet_view.showGridLines = False
head(ws, "UPL — États financiers 2022-2025", "Outil de reconstitution — Université Privée de Libreville — 31 août 2026", 6)
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 112
notes = [
    ("OBJET", "Reconstituer les états financiers demandés par le conducteur d'échange ECOBANK : chiffre d'affaires et résultats sur trois exercices, détail des produits, actifs au 31/12/2025."),
    ("STATUT DES DONNÉES", "2025 : reconstitué à partir du suivi interne (produits × prix publics, charges connues) — calé sur le dossier ECOBANK (CA 95,5 M ; résultat net 17,9 M). 2022 à 2024 : estimatifs, à consolider avec les relevés UGB."),
    ("CELLULES JAUNES", "À saisir ou ajuster (auditeurs réels par promotion, trésorerie du relevé UGB, créances, dettes). Les totaux, l'EBE, l'IS et le résultat net se recalculent seuls."),
    ("ÉQUILIBRE DU BILAN", "Saisissez la trésorerie réelle au 31/12/2025 : le report à nouveau s'ajuste automatiquement pour équilibrer actif = passif (dividendes versés / résultats conservés)."),
    ("IMPÔTS", "Impôt sur les sociétés appliqué à 30 %. Aucune dette financière sur la période. Matériel de faible valeur passé en charges (pas de dotation significative)."),
    ("UTILISATION", "Copier les totaux dans la section « Activités du client » du conducteur (tableau CA / résultat net) et la section « Principaux actifs ». Version imprimable dans le dossier bancaire, section « États financiers 2022-2025 »."),
    ("DIFFUSION", "Document de travail de la direction — à relire avant toute remise extérieure."),
]
r = 4
for k, v in notes:
    ws.cell(r, 2, k).font = F_S
    ws.cell(r, 2).fill = FILL_S
    ws.cell(r, 2).border = THIN
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=2)
    c = ws.cell(r + 1, 2, v); c.font = F_B; c.alignment = LEFT
    ws.row_dimensions[r + 1].height = 42
    r += 3

# ============================================================ COMPTE DE RÉSULTAT
ws = wb.create_sheet("Compte de résultat")
ws.sheet_view.showGridLines = False
head(ws, "Compte de résultat 2022-2025", "2022 à 2024 estimatifs — 2025 reconstitué (FCFA)", 6)
widths = [46, 15, 15, 15, 15, 34]
for j, w_ in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w_
hrow(ws, 4, ["Rubrique", "2022 (est.)", "2023 (est.)", "2024 (est.)", "2025", "Source / note"])

rows = [
    ("Executive MBA — auditeurs (saisie)", 16, 18, 19, 20, "Promotions réelles à confirmer", (1, 2, 3, 4)),
    ("Prix moyen du MBA (FCFA)", 4_000_000, 4_000_000, 4_000_000, 4_000_000, "Grille publique UPL", ()),
    ("CA Executive MBA", "=B5*B6", "=C5*C6", "=D5*D6", "=E5*E6", "Auditeurs × prix", ()),
    ("CA VAE", 1_500_000, 4_000_000, 2_000_000, 7_500_000, "Dossiers suivis", (1, 2, 3, 4)),
    ("CA accompagnement et conseil", 1_000_000, 2_500_000, 1_000_000, 6_000_000, "Prestations", (1, 2, 3, 4)),
    ("CA frais de dossier et inscription", 500_000, 1_500_000, 1_000_000, 2_000_000, "Inscriptions", (1, 2, 3, 4)),
]
r = 5
for lbl, *vals, note, inp in rows:
    row(ws, r, [lbl, *vals, note], inputs=(j + 1 for j in inp))
    r += 1
row(ws, r, ["CHIFFRE D'AFFAIRES", "=SUM(B7:B10)", "=SUM(C7:C10)", "=SUM(D7:D10)", "=SUM(E7:E10)", ""], bold=True, fill=FILL_S)
ca = r
r += 1
rows = [
    ("Enseignants vacataires", -22_000_000, -38_000_000, -36_000_000, -38_000_000, "Vacataires, incl. missions ESSEC Douala", (1, 2, 3, 4)),
    ("Loyer et charges du site", -7_200_000, -7_800_000, -8_400_000, -9_000_000, "Sablière — mensuel", (1, 2, 3, 4)),
    ("Rémunérations et charges sociales", -4_800_000, -5_400_000, -6_000_000, -7_200_000, "Secrétariat", (1, 2, 3, 4)),
    ("Communication", -2_000_000, -3_000_000, -2_500_000, -4_700_000, "Campagnes, supports", (1, 2, 3, 4)),
    ("Déplacements et missions", -1_500_000, -2_000_000, -2_000_000, -2_500_000, "Douala, intérieur", (1, 2, 3, 4)),
    ("Frais généraux, honoraires, assurances", -3_500_000, -5_800_000, -5_100_000, -8_500_000, "", (1, 2, 3, 4)),
    ("Frais d'établissement (lancement)", -14_000_000, 0, 0, 0, "2022 uniquement : agréments, lancement", (1,)),
]
first_ch = r
for lbl, *vals, note, inp in rows:
    row(ws, r, [lbl, *vals, note], inputs=(j + 1 for j in inp))
    r += 1
last_ch = r - 1
row(ws, r, ["TOTAL CHARGES D'EXPLOITATION",
            f"=SUM(B{first_ch}:B{last_ch})", f"=SUM(C{first_ch}:C{last_ch})",
            f"=SUM(D{first_ch}:D{last_ch})", f"=SUM(E{first_ch}:E{last_ch})", ""], bold=True, fill=FILL_S)
r += 1
row(ws, r, ["EXCÉDENT BRUT D'EXPLOITATION",
            f"=B{ca}+B{r-1}", f"=C{ca}+C{r-1}", f"=D{ca}+D{r-1}", f"=E{ca}+E{r-1}",
            "CA − charges (pas de dotation : matériel en charges)"], bold=True, fill=FILL_S)
ebe = r
r += 1
row(ws, r, ["Charges financières", 0, 0, 0, 0, "Aucune dette sur la période"])
r += 1
row(ws, r, ["Impôt sur les sociétés (30 %)",
            f"=-ROUND(B{ebe}*0.3,-5)", f"=-ROUND(C{ebe}*0.3,-5)",
            f"=-ROUND(D{ebe}*0.3,-5)", f"=-ROUND(E{ebe}*0.3,-5)", "Taux 2026"], ())
imp = r
r += 1
row(ws, r, ["RÉSULTAT NET",
            f"=B{ebe}+B{imp}", f"=C{ebe}+C{imp}", f"=D{ebe}+D{imp}", f"=E{ebe}+E{imp}",
            "2025 = 17,9 M (calé sur le dossier ECOBANK)"], bold=True, fill=FILL_S)
rn = r
ws.cell(rn + 2, 1, "Contrôle : 2025 → CA 95 500 000 ; EBE 25 600 000 ; résultat net 17 900 000 (conducteur ECOBANK et annexe financière).").font = F_N
ws.freeze_panes = "B5"

# ============================================================ BILAN 2025
ws = wb.create_sheet("Bilan 2025")
ws.sheet_view.showGridLines = False
head(ws, "Bilan au 31 décembre 2025 (reconstitué)", "Saisir les cellules jaunes d'après le relevé UGB — le report à nouveau équilibre seul (FCFA)", 3)
for j, w_ in enumerate([52, 20, 52], 1):
    ws.column_dimensions[get_column_letter(j)].width = w_
hrow(ws, 4, ["ACTIF", "Montant", "Note"])
r = 5
row(ws, r, ["Aménagements et équipements pédagogiques (brut)", 14_000_000, "Inventaire pédagogique 2022-2025"], inputs=(1,)); r += 1
row(ws, r, ["Amortissements cumulés", -2_100_000, "Linéaire — estimation"], inputs=(1,)); r += 1
row(ws, r, ["Immobilisations nettes", "=B5+B6", ""]); imb = r; r += 1
row(ws, r, ["Créances étudiants", 8_000_000, "≈ 10 % du CA MBA — état des créances joint"], inputs=(1,)); cre = r; r += 1
row(ws, r, ["Trésorerie (compte UGB au 31/12/2025)", 19_500_000, "SAISIR le solde réel du relevé"], inputs=(1,)); tre = r; r += 1
row(ws, r, ["TOTAL ACTIF", f"=B{imb}+B{cre}+B{tre}", ""], bold=True, fill=FILL_S); ta = r
r += 2
hrow(ws, r, ["PASSIF", "Montant", "Note"]); r += 1
row(ws, r, ["Capital social", 20_000_000, "Entièrement libéré — famille 100 %"]); r += 1
row(ws, r, ["Report à nouveau (équilibre automatique)", f"=B{ta}-B{r-1}-B{r+1}-B{r+2}", "Résultats antérieurs distribués / conservés"]); rna = r; r += 1
row(ws, r, ["Résultat net 2025", "='Compte de résultat'!E23", "Lié au compte de résultat"]); r += 1
row(ws, r, ["Dettes fournisseurs et sociales", 1_400_000, "SAISIR le réel"], inputs=(1,)); det = r; r += 1
row(ws, r, ["TOTAL PASSIF", f"=B{rna-1}+B{rna}+B{rna+1}+B{rna+2}", ""], bold=True, fill=FILL_S)
tp = r
r += 2
ws.cell(r, 1, "Contrôle de cohérence : TOTAL ACTIF = TOTAL PASSIF").font = F_S
ws.cell(r + 1, 1, f"=IF(B{ta}=B{tp},\"ÉQUILIBRE — actif = passif\",\"ÉCART — vérifier les saisies\")").font = F_N

# ============================================================ SYNTHÈSE CONDUCTEUR
ws = wb.create_sheet("Synthèse conducteur")
ws.sheet_view.showGridLines = False
head(ws, "Synthèse — sections chiffrées du conducteur ECOBANK", "Valeurs à reporter dans « Activités du client » et « Principaux actifs »", 4)
for j, w_ in enumerate([44, 22, 22, 22], 1):
    ws.column_dimensions[get_column_letter(j)].width = w_
hrow(ws, 4, ["Tableau des trois derniers exercices", "2023", "2024", "2025"])
r = 5
row(ws, r, ["Chiffre d'affaires (FCFA)", "≈ 80 000 000", "≈ 80 000 000", "95 500 000"]); r += 1
row(ws, r, ["Résultat net (FCFA)", "≈ 12 600 000", "≈ 14 000 000", "17 900 000"]); r += 1
row(ws, r, ["Note", "2023-2024 estimatifs — états détaillés en annexe « États financiers 2022-2025 »", "", ""]); r += 2
hrow(ws, r, ["Produits — part du CA 2025", "Part", "Montant (FCFA)", ""]); r += 1
for lbl, p, m in [("Executive MBA (20 auditeurs × 4 000 000)", "83,8 %", "80 000 000"),
                  ("Validation des acquis de l'expérience (VAE)", "7,9 %", "7 500 000"),
                  ("Accompagnement et conseil", "6,3 %", "6 000 000"),
                  ("Frais de dossier et d'inscription", "2,1 %", "2 000 000")]:
    row(ws, r, [lbl, p, m, ""]); r += 1
row(ws, r, ["TOTAL CA RÉCURRENT 2025", "100 %", "95 500 000", ""], bold=True, fill=FILL_S); r += 2
hrow(ws, r, ["Principaux actifs au 31/12/2025", "Valeur", "Nantissement", ""]); r += 1
for lbl, v, n in [("Aménagements et équipements pédagogiques (brut)", "14 000 000", "Non"),
                  ("Créances étudiants", "8 000 000", "Non"),
                  ("Trésorerie (compte UGB)", "19 500 000", "Non"),
                  ("Convention ESSEC de Douala", "Actif immatériel non valorisé", "Non"),
                  ("Bâtiment R+2 en cours (devis ferme)", "219 972 060 TTC", "À constituer")]:
    row(ws, r, [lbl, v, n, ""]); r += 1

# ============================================================ PRÉVISIONS
ws = wb.create_sheet("Prévisions 2026-2030")
ws.sheet_view.showGridLines = False
head(ws, "Prévisions 2026-2030 — dossier ECOBANK", "Hypothèses mesurées ; MBA 20 à 32 auditeurs ; neuf salles livrées sous deux mois", 6)
for j, w_ in enumerate([40, 14, 14, 14, 14, 14], 1):
    ws.column_dimensions[get_column_letter(j)].width = w_
hrow(ws, 4, ["Indicateur", "2026", "2027", "2028", "2029", "2030"])
data = [("Effectifs UPL", "80", "120", "170", "210", "230"),
        ("Chiffre d'affaires (M FCFA)", "119,5", "164,0", "206,9", "242,8", "252,8"),
        ("EBE (M FCFA)", "50,0", "79,0", "106,0", "127,0", "132,0"),
        ("Dotations (M FCFA)", "10,0", "17,5", "17,5", "17,5", "17,5"),
        ("Résultat net (M FCFA)", "27,5", "35,0", "58,0", "76,0", "82,5"),
        ("CAF (M FCFA)", "33,5", "42,3", "62,1", "78,1", "83,2"),
        ("Service de la dette (M FCFA)", "6,5", "30,5", "43,9", "43,9", "43,9"),
        ("DSCR (CAF / dette)", "5,15 x", "1,39 x", "1,41 x", "1,78 x", "1,90 x")]
r = 5
for line in data:
    row(ws, r, list(line)); r += 1
r += 1
ws.cell(r, 1, "Crédit : 260 000 000 FCFA — 120 mois dont 12 de différé de capital — taux indicatif 10 % — ").font = F_S
ws.cell(r + 1, 1, "mensualité en différé 2 166 667 FCFA (intérêts seuls) ; en régime 3 660 458 FCFA ; coût total ≈ 421,3 M FCFA.").font = F_S

wb.save(OUT)
print("OK :", OUT)
